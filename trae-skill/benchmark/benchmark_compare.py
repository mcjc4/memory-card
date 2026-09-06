#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
错题记忆卡横评对比报告生成器
读取两个模型的 xlsx（7 列同结构），输出对比报告。

用法:
  python benchmark_compare.py deepseek.xlsx glm53.xlsx > report.md
  python benchmark_compare.py deepseek.xlsx  # 单模型评估
"""
import sys
from pathlib import Path
import openpyxl


def load_cards(xlsx_path):
    """读取 xlsx，返回 (subjects, cards_list)"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    cards = []
    subjects = {}
    for r in range(2, ws.max_row + 1):
        row = {
            'subject': ws.cell(r, 1).value,
            'chapter': ws.cell(r, 2).value,
            'question': ws.cell(r, 3).value,
            'signal': ws.cell(r, 4).value,
            'conclusion': ws.cell(r, 5).value,
            'source': ws.cell(r, 6).value,
            'link': ws.cell(r, 7).value,
        }
        cards.append(row)
        s = row['subject'] or '未知'
        subjects[s] = subjects.get(s, 0) + 1
    return subjects, cards


def per_question_stats(cards):
    """统计每题拆几条信号卡（按 link 分组）"""
    stats = {}
    for c in cards:
        link = c.get('link') or c.get('source') or '?'
        stats[link] = stats.get(link, 0) + 1
    return stats


def latex_quality(cards):
    """检测 LaTeX 合规性：$...$ 是否成对、是否有遗留 ${、公式是否合理"""
    bad = 0
    samples = []
    for i, c in enumerate(cards):
        text = (c.get('signal') or '') + (c.get('conclusion') or '')
        dollar_count = text.count('$')
        if dollar_count % 2 != 0:
            bad += 1
            samples.append((i + 2, '奇数 $', text[:60]))
        # 检查常见 LaTeX 错误：\\{ 未闭合
        if '\\{' in text and '\\}' not in text:
            bad += 1
            samples.append((i + 2, '未闭合 \\{', text[:60]))
        if '\\}' in text and '\\{' not in text:
            bad += 1
            samples.append((i + 2, '未匹配 \\}', text[:60]))
    return bad, samples


def generalization_score(cards):
    """通用化自检：检测结论里是否残留本题具体数字（如本题年份、试卷号）"""
    suspicious = []
    for i, c in enumerate(cards):
        concl = c.get('conclusion') or ''
        # 检测：具体年份（2023、2024）+ 试卷号（如「24」「25」）出现 → 可能就题论题
        import re
        # 模式：2024-XX、25届、某具体分数（如 a=0.4、v=12）
        patterns = [
            (r'\d+\.\d+\s*m/s', '具体速度值'),
            (r'\d+°', '具体角度'),
            (r'2024|2023|2025', '具体年份'),
        ]
        for pat, label in patterns:
            if re.search(pat, concl) and label in ('具体年份',):
                # 年份可能出现在 source 列，conclusion 不应该有
                if not c.get('source') or str(re.search(pat, concl).group()) not in str(c.get('source')):
                    suspicious.append((i + 2, label, concl[:80]))
                    break
    return len(suspicious), suspicious


def report_single(xlsx_path):
    p = Path(xlsx_path)
    print(f"# 📊 {p.stem} 评估报告")
    print()
    subjects, cards = load_cards(xlsx_path)
    total = len(cards)
    print(f"**文件**：`{p.name}`")
    print(f"**总卡片数**：{total}")
    print()
    print("## 科目分布")
    for s, n in sorted(subjects.items(), key=lambda x: -x[1]):
        print(f"- {s}: {n} 条")
    print()
    per_q = per_question_stats(cards)
    print(f"## 每题拆解粒度")
    print(f"- 共 {len(per_q)} 个不同题源")
    print(f"- 平均 {total / max(len(per_q), 1):.2f} 条/题")
    print(f"- 分布：{dict(sorted(per_q.items(), key=lambda x: x[1]))}")
    print()
    bad, samples = latex_quality(cards)
    print(f"## LaTeX 合规性")
    print(f"- 不合规卡片数：{bad}/{total}")
    if samples:
        print(f"- 样例（前 3）：")
        for row, issue, txt in samples[:3]:
            print(f"  - 行 {row} [{issue}]: {txt}")
    print()
    sus, samples = generalization_score(cards)
    print(f"## 通用化自检")
    print(f"- 疑似「就题论题」数：{sus}/{total}")
    if samples:
        print(f"- 样例（前 3）：")
        for row, label, txt in samples[:3]:
            print(f"  - 行 {row} [{label}]: {txt}")
    print()


def report_compare(path_a, path_b):
    pa = Path(path_a)
    pb = Path(path_b)
    model_a = pa.stem.replace('错题本5题_记忆卡_物理_', '').replace('_导入版_LaTeX公式', '').replace('_纯文本版', '')
    model_b = pb.stem.replace('错题本5题_记忆卡_物理_', '').replace('_导入版_LaTeX公式', '').replace('_纯文本版', '')

    sub_a, cards_a = load_cards(pa)
    sub_b, cards_b = load_cards(pb)

    print(f"# 🆚 横评对比报告：{model_a} vs {model_b}")
    print()
    print(f"- 📄 {pa.name}（{len(cards_a)} 条）")
    print(f"- 📄 {pb.name}（{len(cards_b)} 条）")
    print()

    # 总数对比
    print("## 卡片总数")
    print(f"| 模型 | 卡片数 | 差值 |")
    print(f"|:--|:--:|:--:|")
    diff = len(cards_b) - len(cards_a)
    print(f"| {model_a} | {len(cards_a)} | — |")
    print(f"| {model_b} | {len(cards_b)} | {diff:+d} |")
    print()

    # 每题拆解
    per_a = per_question_stats(cards_a)
    per_b = per_question_stats(cards_b)
    print("## 每题拆解粒度")
    print(f"| 题链接 | {model_a} | {model_b} | 差值 |")
    print(f"|:--|:--:|:--:|:--:|")
    all_links = sorted(set(per_a) | set(per_b))
    for link in all_links:
        a = per_a.get(link, 0)
        b = per_b.get(link, 0)
        diff = b - a
        link_short = link.replace('https://zujuan.xkw.com/', '') if link else '?'
        print(f"| {link_short} | {a} | {b} | {diff:+d} |")
    print()

    # LaTeX 合规性
    bad_a, _ = latex_quality(cards_a)
    bad_b, _ = latex_quality(cards_b)
    print("## LaTeX 合规性")
    print(f"| 模型 | 不合规数 | 合规率 |")
    print(f"|:--|:--:|:--:|")
    total_a = len(cards_a)
    total_b = len(cards_b)
    rate_a = f"{(total_a - bad_a) / max(total_a, 1) * 100:.1f}%"
    rate_b = f"{(total_b - bad_b) / max(total_b, 1) * 100:.1f}%"
    print(f"| {model_a} | {bad_a}/{total_a} | {rate_a} |")
    print(f"| {model_b} | {bad_b}/{total_b} | {rate_b} |")
    print()

    # 通用化
    sus_a, _ = generalization_score(cards_a)
    sus_b, _ = generalization_score(cards_b)
    print("## 通用化自检")
    print(f"| 模型 | 疑似就题论题 | 通用化率 |")
    print(f"|:--|:--:|:--:|")
    rate_a = f"{(total_a - sus_a) / max(total_a, 1) * 100:.1f}%"
    rate_b = f"{(total_b - sus_b) / max(total_b, 1) * 100:.1f}%"
    print(f"| {model_a} | {sus_a}/{total_a} | {rate_a} |")
    print(f"| {model_b} | {sus_b}/{total_b} | {rate_b} |")
    print()

    # 结论：哪一方更适合「记忆卡应用」？
    score_a = (total_a - bad_a - sus_a) / max(total_a, 1) * 100
    score_b = (total_b - bad_b - sus_b) / max(total_b, 1) * 100
    print("## 综合评分")
    print(f"- **{model_a}**：{score_a:.1f} 分（合规 + 通用化）")
    print(f"- **{model_b}**：{score_b:.1f} 分")
    winner = model_a if score_a > score_b else (model_b if score_b > score_a else '平局')
    print(f"- 🏆 本次胜者：**{winner}**")
    print()

    # 建议
    print("## 改进建议")
    if bad_a > 0 or bad_b > 0:
        print("- 修正 LaTeX 符号对（`$` 配对、`\\{`/`\\}` 闭合）")
    if sus_a > 0 or sus_b > 0:
        print("- 进一步剥离具体数字，结论做题型/模式级抽象")
    if abs(len(cards_a) - len(cards_b)) > 2:
        print(f"- 拆解粒度差异较大（{abs(len(cards_a) - len(cards_b))} 条），建议两模型按同一题目清单各自拆解后比对")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法:")
        print("  python benchmark_compare.py <xlsx>                 # 单模型评估")
        print("  python benchmark_compare.py <model_a.xlsx> <model_b.xlsx>   # 横评对比")
        sys.exit(1)

    if len(sys.argv) == 2:
        report_single(sys.argv[1])
    else:
        report_compare(sys.argv[1], sys.argv[2])