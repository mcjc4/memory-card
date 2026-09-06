#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
wrongbook.py — 错题整理 → 双版本 xlsx 生成器（强制双版本输出）

【硬约束】每次必须同时生成两个文件，缺一不可：
  1. <name>_纯文本版.xlsx          （公式用 Unicode 上下标：v²、√、α）
  2. <name>_导入版_LaTeX公式.xlsx  （公式用 $...$ 定界：$v^2$、$\\sqrt{}$）

用法：
  # 1. 打印空模板（仿造后填入题目内容）
  python wrongbook.py template

  # 2. 从 JSON 构建双版本 xlsx
  python wrongbook.py build cards.json ./output [model_name]

  # 3. 当作 Python 模块 import
  from wrongbook import build_both
  build_both(cards, Path('./output/错题本5题_记忆卡_物理'), 'DeepSeek')

设计原则：
  - 物理上不可能只生成一个：build_both 内置 3 道断言
  - 命名规则遵循「错题本XX题_记忆卡_<科目>_<模型名>.xlsx」便于多模型横评归档
  - LaTeX 版公式标记与纯文本版完全相同的内容，仅做公式形态变换
  - 不依赖任何在线资源，离线运行

Author: 小宇（错题整理专家）
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============== 常量 ==============
HEADERS = ['科目', '章节知识点', '原题题干', '题干信号', '结论/易错点', '出处', '题目链接']
COL_WIDTHS = [8, 22, 52, 46, 52, 20, 42]
HEADER_FILL_HEX = '185FA5'  # 深蓝表头
PLAIN_SUFFIX = '_纯文本版.xlsx'
LATEX_SUFFIX = '_导入版_LaTeX公式.xlsx'


# ============== 公式变换 ==============
def latex_to_plain(text: str) -> str:
    """
    把 LaTeX 公式文本转换为 Unicode 纯文本形式（用于纯文本版）。
    支持：
      - $...$ 区间（内层 LaTeX 转 Unicode 后去掉 $）
      - 散落的 LaTeX 标记（无 $ 包裹）
      - 常见希腊字母与符号
    """
    if not isinstance(text, str):
        return text

    s = text
    # 先处理 $...$ 区间：内层做完整数学变换
    def _dollar_cb(m):
        return _transform_math(m.group(1))

    s = re.sub(r'\$([^$]+)\$', _dollar_cb, s)
    # 再处理剩余的散落 LaTeX（无 $ 包裹）
    s = _transform_math(s)
    return s


def _transform_math(s: str) -> str:
    """对单段文本做一次完整 LaTeX → Unicode 数学变换"""
    # 上下标花括号形式
    s = re.sub(r'\^\{([^}]+)\}', lambda m: _super_digits(m.group(1)), s)
    s = re.sub(r'_\{([^}]+)\}', lambda m: _sub_chars(m.group(1)), s)
    # 上下标单字符
    s = re.sub(r'\^([0-9a-zA-Z\+\-])', lambda m: _super_one(m.group(1)), s)
    s = re.sub(r'_([0-9a-zA-Z\+\-])', lambda m: _sub_one(m.group(1)), s)

    # \frac{a}{b} → (a/b)
    s = re.sub(r'\\d?frac\{([^{}]+)\}\{([^{}]+)\}', r'(\1/\2)', s)

    # \sqrt[n]{x} → ⁿ√x，\sqrt{x} → √x
    s = re.sub(r'\\sqrt\[([^\]]+)\]\{([^{}]+)\}', r'^\1√\2', s)
    s = re.sub(r'\\sqrt\{([^{}]+)\}', r'√\1', s)

    # 希腊字母 + 符号
    for k, v in _greek_map().items():
        s = s.replace(k, v)
    for k, v in _sym_map().items():
        s = s.replace(k, v)

    # 清理转义与花括号
    s = s.replace('\\{', '{').replace('\\}', '}')
    s = s.replace('\\(', '(').replace('\\)', ')')
    s = s.replace('\\,', '').replace('\\;', ' ').replace('\\!', '')
    s = s.replace('{', '').replace('}', '')
    # 合并「希腊字母 + 空格 + 字母」成紧凑形式（Δ x → Δx）
    s = re.sub(r'([αβγδεζηθικλμνξοπρστυφχψωΑΒΓΔΕΖΗΘΙΚΛΜΝΞΟΠΡΣΤΥΦΧΨΩ])\s+([a-zA-Z])', r'\1\2', s)

    return s


def _greek_map() -> dict:
    return {
        r'\alpha': 'α', r'\beta': 'β', r'\gamma': 'γ', r'\delta': 'δ',
        r'\epsilon': 'ε', r'\varepsilon': 'ε', r'\zeta': 'ζ', r'\eta': 'η',
        r'\theta': 'θ', r'\vartheta': 'ϑ', r'\iota': 'ι', r'\kappa': 'κ',
        r'\lambda': 'λ', r'\mu': 'μ', r'\nu': 'ν', r'\xi': 'ξ',
        r'\pi': 'π', r'\varpi': 'ϖ', r'\rho': 'ρ', r'\varrho': 'ϱ',
        r'\sigma': 'σ', r'\varsigma': 'ς', r'\tau': 'τ', r'\upsilon': 'υ',
        r'\phi': 'φ', r'\varphi': 'ϕ', r'\chi': 'χ', r'\psi': 'ψ',
        r'\omega': 'ω',
        r'\Gamma': 'Γ', r'\Delta': 'Δ', r'\Theta': 'Θ', r'\Lambda': 'Λ',
        r'\Xi': 'Ξ', r'\Pi': 'Π', r'\Sigma': 'Σ', r'\Upsilon': 'Υ',
        r'\Phi': 'Φ', r'\Psi': 'Ψ', r'\Omega': 'Ω',
    }


def _sym_map() -> dict:
    return {
        r'\leq': '≤', r'\le': '≤', r'\geq': '≥', r'\ge': '≥',
        r'\neq': '≠', r'\ne': '≠', r'\approx': '≈', r'\sim': '~',
        r'\times': '×', r'\div': '÷', r'\pm': '±', r'\mp': '∓',
        r'\to': '→', r'\rightarrow': '→', r'\leftarrow': '←',
        r'\Rightarrow': '⇒', r'\Leftarrow': '⇐', r'\Leftrightarrow': '⇔',
        r'\perp': '⊥', r'\parallel': '∥', r'\angle': '∠',
        r'\triangle': '△', r'\cap': '∩', r'\cup': '∪',
        r'\subset': '⊂', r'\supset': '⊃', r'\subseteq': '⊆', r'\supseteq': '⊇',
        r'\in': '∈', r'\notin': '∉', r'\varnothing': '∅',
        r'\infty': '∞', r'\partial': '∂', r'\nabla': '∇',
        r'\sum': '∑', r'\prod': '∏', r'\int': '∫',
        r'\cdot': '·', r'\cdots': '⋯', r'\ldots': '…', r'\dots': '…',
        r'\quad': '  ', r'\,': '', r'\;': ' ',
        r'\ln': 'ln', r'\log': 'log',
        r'\sin': 'sin', r'\cos': 'cos', r'\tan': 'tan',
        r'\mathbb{R}': 'ℝ', r'\mathbb{N}': 'ℕ', r'\mathbb{Z}': 'ℤ', r'\mathbb{Q}': 'ℚ',
        r'\R': 'ℝ', r'\N': 'ℕ', r'\Z': 'ℤ',
    }


def _super_one(s: str) -> str:
    """把单字符转上标 Unicode（找不到映射则保留原字符）"""
    sup = {
        '0': '⁰', '1': '¹', '2': '²', '3': '³', '4': '⁴',
        '5': '⁵', '6': '⁶', '7': '⁷', '8': '⁸', '9': '⁹',
        'a': 'ᵃ', 'b': 'ᵇ', 'c': 'ᶜ', 'd': 'ᵈ', 'e': 'ᵉ',
        'f': 'ᶠ', 'g': 'ᵍ', 'h': 'ʰ', 'i': 'ⁱ', 'j': 'ʲ',
        'k': 'ᵏ', 'l': 'ˡ', 'm': 'ᵐ', 'n': 'ⁿ', 'o': 'ᵒ',
        'p': 'ᵖ', 'r': 'ʳ', 's': 'ˢ', 't': 'ᵗ', 'u': 'ᵘ',
        'v': 'ᵛ', 'w': 'ʷ', 'x': 'ˣ', 'y': 'ʸ', 'z': 'ᶻ',
        '+': '⁺', '-': '⁻',
    }
    return sup.get(s, s)


def _sub_one(s: str) -> str:
    """把单字符转下标 Unicode（找不到映射则保留原字符）"""
    sub = {
        '0': '₀', '1': '₁', '2': '₂', '3': '₃', '4': '₄',
        '5': '₅', '6': '₆', '7': '₇', '8': '₈', '9': '₉',
        'a': 'ₐ', 'e': 'ₑ', 'h': 'ₕ', 'i': 'ᵢ', 'j': 'ⱼ',
        'k': 'ₖ', 'l': 'ₗ', 'm': 'ₘ', 'n': 'ₙ', 'o': 'ₒ',
        'p': 'ₚ', 'r': 'ᵣ', 's': 'ₛ', 't': 'ₜ', 'u': 'ᵤ',
        'v': 'ᵥ', 'x': 'ₓ',
    }
    return sub.get(s, s)


def _super_digits(s: str) -> str:
    """把字符串转上标 Unicode"""
    return ''.join(_super_one(c) for c in s)


def _sub_chars(s: str) -> str:
    """把字符串转下标 Unicode"""
    return ''.join(_sub_one(c) for c in s)


# ============== 样式 ==============
def _style_header(ws):
    fill = PatternFill('solid', fgColor=HEADER_FILL_HEX)
    font = Font(bold=True, color='FFFFFF', size=11)
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, col, h)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.row_dimensions[1].height = 26


def _style_data(ws, n_rows):
    thin = Side(border_style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    align_center = Alignment(horizontal='center', vertical='top', wrap_text=True)
    link_font = Font(color='0563C1', underline='single', size=11)
    for r in range(2, n_rows + 2):
        ws.row_dimensions[r].height = 26
        for c in range(1, 8):
            cell = ws.cell(r, c)
            cell.alignment = align_center if c == 1 else align_left
            cell.border = border
        link_cell = ws.cell(r, 7)
        if link_cell.value:
            link_cell.hyperlink = link_cell.value
            link_cell.font = link_font


def _apply_columns(ws):
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(7)}{ws.max_row}'


# ============== 核心写文件 ==============
def write_xlsx(cards: List[Dict[str, Any]], out_path: Path, latex_mode: bool = False) -> Path:
    """
    把 cards 列表写入 xlsx 文件。
    :param latex_mode: True = 保持 LaTeX 公式；False = 转换为 Unicode 纯文本
    """
    wb = Workbook()
    ws = wb.active
    ws.title = '记忆卡'
    _style_header(ws)

    for r, card in enumerate(cards, 2):
        ws.cell(r, 1, card.get('subject', ''))
        ws.cell(r, 2, card.get('chapter', ''))
        question = card.get('question', '')
        signal = card.get('signal', '')
        conclusion = card.get('conclusion', '')
        if not latex_mode:
            question = latex_to_plain(question)
            signal = latex_to_plain(signal)
            conclusion = latex_to_plain(conclusion)
        ws.cell(r, 3, question)
        ws.cell(r, 4, signal)
        ws.cell(r, 5, conclusion)
        ws.cell(r, 6, card.get('source', ''))
        ws.cell(r, 7, card.get('link', ''))

    _style_data(ws, len(cards))
    _apply_columns(ws)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


# ============== 双版本强制输出 ==============
def build_both(cards: List[Dict[str, Any]], base_path: Path, model_name: str = 'Unknown') -> Dict[str, Path]:
    """
    【核心入口】强制生成双版本 xlsx，缺一不可。

    :param cards: 卡片列表，每条 dict 包含 subject/chapter/question/signal/conclusion/source/link
    :param base_path: 基础路径（不含后缀），例如 Path('./output/错题本5题_记忆卡_物理_DeepSeek')
    :param model_name: 模型名（用于跨模型横评归档），若已在 base_path 中则不重复追加
    :return: {'plain': Path, 'latex': Path}
    """
    if not cards:
        raise ValueError('cards 不能为空')

    base_path = Path(base_path)
    stem = base_path.stem
    # 自动追加模型名（防止横评归档冲突）
    if model_name and model_name != 'Unknown' and model_name not in stem:
        stem = f'{stem}_{model_name}'
    base_dir = base_path.parent

    plain_path = base_dir / f'{stem}{PLAIN_SUFFIX}'
    latex_path = base_dir / f'{stem}{LATEX_SUFFIX}'

    write_xlsx(cards, plain_path, latex_mode=False)
    write_xlsx(cards, latex_path, latex_mode=True)

    # ========== 强制断言：物理上不可能只生成一个 ==========
    assert plain_path.exists(), f'❌ 纯文本版未生成: {plain_path}'
    assert latex_path.exists(), f'❌ LaTeX 版未生成: {latex_path}'
    assert plain_path.stat().st_size > 1000, f'❌ 纯文本版文件异常小: {plain_path.stat().st_size}B'
    assert latex_path.stat().st_size > 1000, f'❌ LaTeX 版文件异常小: {latex_path.stat().st_size}B'

    return {'plain': plain_path, 'latex': latex_path}


# ============== CLI ==============
def _print_template():
    """打印空模板 JSON 供仿造"""
    template = [
        {
            'subject': '物理',
            'chapter': '运动学·连接体',
            'question': '题1（2013·新课标I）：两玩具车 A、B 用橡皮筋连接，橡皮筋从原长 l 伸长到 3l，A 沿 y 轴匀加速...',
            'signal': '看到【连接体上的标记点，且题设"伸长均匀"】',
            'conclusion': '想到【标记点始终按原长比例分割两端点连线，可用相似三角形建立几何约束】；警惕【把"伸长均匀"误解为长度不变】',
            'source': '2013·新课标I',
            'link': 'https://zujuan.xkw.com/13q1685290.html',
        },
        {
            'subject': '物理',
            'chapter': '运动学·匀变速',
            'question': '题2：质点做匀减速直线运动，末速度为 0...',
            'signal': '看到【末速度为 0 的匀减速】',
            'conclusion': '想到【用逆向思维当作"初速为 0 匀加速"，时间比 $1:(\\sqrt{2}-1):(\\sqrt{3}-\\sqrt{2}):\\dots$】；警惕【正向代入常规公式导致复杂度翻倍】',
            'source': '23-24高三上·广东深圳',
            'link': 'https://zujuan.xkw.com/13q20505949.html',
        },
    ]
    print(json.dumps(template, ensure_ascii=False, indent=2))
    print()
    print('# ============================================================')
    print('# 用法：复制上方 JSON 到 cards.json，每条卡片包含 7 个字段：')
    print('#   subject    科目（物理/化学/英语/政治/数学/语文...）')
    print('#   chapter    章节知识点')
    print('#   question   原题题干（题N：…）')
    print('#   signal     题干信号（看到【X】）')
    print('#   conclusion 结论/易错点（想到【Y】；警惕【Z】）')
    print('#   source     出处')
    print('#   link       题目详情链接')
    print('# 然后执行：')
    print('#   python wrongbook.py build cards.json ./output DeepSeek')
    print('# ============================================================')


def _cli():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    cmd = sys.argv[1]
    if cmd == 'template':
        _print_template()
    elif cmd == 'build':
        if len(sys.argv) < 4:
            print('用法: python wrongbook.py build <cards.json> <out_dir> [model_name]')
            sys.exit(1)
        cards_path = Path(sys.argv[2])
        out_dir = Path(sys.argv[3])
        model_name = sys.argv[4] if len(sys.argv) > 4 else 'Unknown'
        cards = json.loads(cards_path.read_text(encoding='utf-8'))
        # 自动统计题数（按 source 去重）
        n_questions = len(set(c.get('source', '') for c in cards))
        subject = cards[0].get('subject', '通用') if cards else '通用'
        base = out_dir / f'错题本{n_questions}题_记忆卡_{subject}'
        result = build_both(cards, base, model_name)
        print('✅ 双版本已生成（强制保证）：')
        print(f'  📄 纯文本版:   {result["plain"]}')
        print(f'  📐 LaTeX 版:    {result["latex"]}')
        print(f'  📊 共 {len(cards)} 条信号卡，来自 {n_questions} 道题')
    else:
        print(f'未知命令: {cmd}')
        print('可用命令: template / build')
        sys.exit(1)


if __name__ == '__main__':
    _cli()
